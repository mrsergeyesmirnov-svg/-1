from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .models import Payment

AMOUNT_KEYS = (
    "сумма к оплате",
    "сумма платежа",
    "сумма оплаты",
    "сумма документа",
    "сумма с ндс",
    "сумма без ндс",
    "сумма руб",
    "сумма, руб",
    "сумма в руб",
    "сумма",
    "к оплате",
    "итого к оплате",
    "итого",
    "всего",
    "amount",
    "summa",
    "sum",
    "value",
    "оплата",
    "платёж",
    "платеж",
    "стоимость",
    "оборот",
    "расход",
    "списание",
    "руб",
    "rur",
    "rub",
)

DEBIT_KEYS = ("дебет", "дебит", "списано", "расход", "outcome", "debit")
CREDIT_KEYS = ("кредит", "зачислено", "приход", "income", "credit")

DATE_KEYS = (
    "дата оплаты",
    "дата платежа",
    "дата документа",
    "дата операции",
    "дата проводки",
    "плановая дата",
    "дата",
    "срок оплаты",
    "срок",
    "день",
    "date",
    "pay date",
    "due",
    "operation date",
)

TITLE_KEYS = (
    "назначение платежа",
    "назначение",
    "описание",
    "статья ддс",
    "статья",
    "основание",
    "title",
    "payment purpose",
    "purpose",
    "платёжка",
    "платежка",
    "содержание",
    "операция",
    "за что",
)

COUNTERPARTY_KEYS = (
    "контрагент",
    "поставщик",
    "получатель",
    "vendor",
    "partner",
    "организация контрагент",
    "корреспондент",
    "кому",
    "название контрагента",
)

# Наше юрлицо / р/с, с которого платим
ACCOUNT_KEYS = (
    "расчётный счёт",
    "расчетный счет",
    "р/с",
    "банк. счет",
    "банковский счет",
    "счёт списания",
    "счет списания",
    "счёт",
    "счет",
    "account",
    "наше юрлицо",
    "наша организация",
    "организация",
    "юрлицо",
    "юр. лицо",
    "юр лицо",
    "плательщик",
    "компания",
    "точка",
    "филиал",
    "бренд",
)

NOTE_KEYS = (
    "примечание",
    "комментарий",
    "коммент",
    "note",
    "comment",
    "пояснение",
    "заметка",
)

STATUS_KEYS = ("статус", "status", "состояние", "оплачено", "признак")

ORG_HINT = re.compile(r"\b(ооо|оао|ао|зао|ип|пао|тоо|llc|ltd)\b", re.I)


def _norm(v: Any) -> str:
    s = str(v or "").replace("\n", " ").replace("\r", " ")
    s = s.replace("ё", "е")
    return re.sub(r"\s+", " ", s).strip().lower()


def _find_col(headers: list[str], keys: tuple[str, ...], used: set[int] | None = None) -> int | None:
    used = used or set()
    norms = [_norm(h) for h in headers]
    for key in sorted(keys, key=len, reverse=True):
        kn = _norm(key)
        if not kn:
            continue
        for i, hn in enumerate(norms):
            if i in used or not hn:
                continue
            if hn == kn:
                return i
            # Short keys (in, sum…) — only exact match, avoid «N»⊂«in»
            if len(kn) <= 3:
                continue
            if kn in hn:
                return i
            # Header shorter than key only if almost the same token
            if len(hn) >= 4 and hn in kn:
                return i
    return None


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            d = from_excel(v)
            if isinstance(d, datetime):
                return d.date()
            if isinstance(d, date):
                return d
        except Exception:
            return None
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y", "%Y.%m.%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    m = re.match(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _parse_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    s = s.replace("\u00a0", " ").replace(" ", "")
    s = s.replace("₽", "").replace("руб.", "").replace("руб", "").replace("RUB", "").replace("rub", "")
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in {".", "-", "-."}:
        return None
    try:
        return abs(float(s))
    except ValueError:
        return None


def _score_amount_col(rows: list[tuple], col: int, start: int) -> int:
    ok = 0
    for row in rows[start : start + 80]:
        if col >= len(row):
            continue
        if _parse_amount(row[col]) is not None:
            ok += 1
    return ok


def _score_date_col(rows: list[tuple], col: int, start: int) -> int:
    ok = 0
    for row in rows[start : start + 80]:
        if col >= len(row):
            continue
        if _parse_date(row[col]) is not None:
            ok += 1
    return ok


def _text_stats(rows: list[tuple], col: int, start: int) -> tuple[int, int, int, float]:
    """nonempty, unique, org_hits, avg_len"""
    vals = []
    org_hits = 0
    for row in rows[start : start + 80]:
        if col >= len(row) or row[col] is None:
            continue
        # skip pure numbers/dates
        if _parse_amount(row[col]) is not None and not re.search(r"[a-zа-я]", _norm(row[col])):
            continue
        if _parse_date(row[col]) is not None and len(str(row[col]).strip()) <= 12:
            continue
        s = str(row[col]).strip()
        if not s:
            continue
        vals.append(s)
        if ORG_HINT.search(s):
            org_hits += 1
    unique = len(set(vals))
    avg_len = (sum(len(v) for v in vals) / len(vals)) if vals else 0.0
    return len(vals), unique, org_hits, avg_len


def _detect_header(rows: list[tuple]) -> tuple[int, list[str]]:
    """Return (header_row_index, headers). header_idx=-1 means no header row."""
    best_idx = 0
    best_cells = [str(c or f"col{i}") for i, c in enumerate(rows[0] or ())]
    best_score = -1
    for i, row in enumerate(rows[:30]):
        if not row:
            continue
        cells = [str(c or "") for c in row]
        joined = " ".join(_norm(c) for c in cells)
        score = 0
        for token in (
            "сумма",
            "amount",
            "дата",
            "date",
            "контрагент",
            "поставщик",
            "получатель",
            "назначение",
            "оплат",
            "счет",
            "счёт",
            "дебет",
            "кредит",
            "расход",
            "приход",
            "организация",
            "юрлицо",
            "коммент",
            "кому",
            "за что",
        ):
            if token in joined:
                score += 4
        numericish = sum(
            1 for c in cells if _parse_amount(c) is not None and not re.search(r"[a-zа-я]", _norm(c))
        )
        dateish = sum(1 for c in cells if _parse_date(c) is not None)
        # Data row pretending to be header
        if numericish >= 1 and dateish >= 1:
            score -= 8
        if numericish >= max(2, len(cells) // 2):
            score -= 4
        if score > 0:
            for j, _ in enumerate(cells):
                score += min(4, _score_amount_col(rows, j, i + 1) // 4)
        if score > best_score:
            best_score = score
            best_idx = i
            best_cells = cells

    if best_score < 6:
        width = max(len(r) for r in rows[:20]) if rows else 0
        return -1, [f"col{i}" for i in range(width)]
    return best_idx, best_cells


def _resolve_amount_cols(
    headers: list[str], rows: list[tuple], header_idx: int
) -> tuple[int | None, int | None, int | None]:
    amount = _find_col(headers, AMOUNT_KEYS)
    debit = _find_col(headers, DEBIT_KEYS)
    credit = _find_col(headers, CREDIT_KEYS)

    if amount is not None and debit is not None and amount == debit:
        amount = None
    if amount is not None and credit is not None and amount == credit:
        amount = None

    if amount is None and debit is None and credit is None:
        start = header_idx + 1
        width = max((len(r) for r in rows[start : start + 40]), default=0) or len(headers)
        scored = [(_score_amount_col(rows, c, start), c) for c in range(width)]
        scored.sort(reverse=True)
        if scored and scored[0][0] >= 2:
            amount = scored[0][1]

    return amount, debit, credit


def _infer_cols(
    headers: list[str],
    rows: list[tuple],
    header_idx: int,
    reserved: set[int],
    amount_col: int | None = None,
) -> dict[str, int | None]:
    cols: dict[str, int | None] = {
        "date": _find_col(headers, DATE_KEYS, reserved),
        "title": _find_col(headers, TITLE_KEYS, reserved),
        "counterparty": _find_col(headers, COUNTERPARTY_KEYS, reserved),
        "account": _find_col(headers, ACCOUNT_KEYS, reserved),
        "status": _find_col(headers, STATUS_KEYS, reserved),
        "note": _find_col(headers, NOTE_KEYS, reserved),
    }
    used = set(reserved)
    for v in cols.values():
        if v is not None:
            used.add(v)

    start = header_idx + 1
    width = max((len(r) for r in rows[start : start + 50]), default=0) or len(headers)

    if cols["date"] is None:
        scored = [(_score_date_col(rows, c, start), c) for c in range(width) if c not in used]
        scored.sort(reverse=True)
        min_hits = 2 if header_idx < 0 else 3
        if scored and scored[0][0] >= min_hits:
            cols["date"] = scored[0][1]
            used.add(scored[0][1])

    text_rank = []
    for c in range(width):
        if c in used:
            continue
        nonempty, unique, org_hits, avg_len = _text_stats(rows, c, start)
        if nonempty < 1:
            continue
        text_rank.append(
            {
                "c": c,
                "nonempty": nonempty,
                "unique": unique,
                "org": org_hits,
                "avg": avg_len,
                "header": _norm(headers[c]) if c < len(headers) else "",
                "left_of_amount": amount_col is not None and c < amount_col,
                "right_of_amount": amount_col is not None and c > amount_col,
            }
        )

    def take(pred) -> int | None:
        nonlocal text_rank
        for i, item in enumerate(text_rank):
            if pred(item):
                text_rank.pop(i)
                used.add(item["c"])
                return item["c"]
        return None

    # Header-based leftovers first
    if cols["account"] is None:
        cols["account"] = take(
            lambda x: any(k in x["header"] for k in ("организ", "юр", "счет", "счёт", "плател", "компан", "филиал", "наше"))
        )
    if cols["counterparty"] is None:
        cols["counterparty"] = take(
            lambda x: any(k in x["header"] for k in ("контраг", "постав", "получ", "кому", "vendor"))
        )
    if cols["title"] is None:
        cols["title"] = take(
            lambda x: any(k in x["header"] for k in ("назнач", "описан", "стать", "за что", "purpose", "содерж"))
        )
    if cols["note"] is None:
        cols["note"] = take(lambda x: any(k in x["header"] for k in ("коммент", "примеч", "note", "поясн")))

    # Position heuristic: left of amount = кому/за что; right = наше юрлицо/коммент
    if amount_col is not None:
        if cols["counterparty"] is None:
            cols["counterparty"] = take(lambda x: x["left_of_amount"] and x["org"] >= 1)
        if cols["counterparty"] is None:
            cols["counterparty"] = take(lambda x: x["left_of_amount"] and x["unique"] >= 1 and x["avg"] <= 80)
        if cols["title"] is None:
            cols["title"] = take(lambda x: x["left_of_amount"] and x["avg"] >= 8)
        if cols["account"] is None:
            cols["account"] = take(lambda x: x["right_of_amount"] and (x["org"] >= 1 or x["avg"] <= 50))
        if cols["note"] is None:
            cols["note"] = take(lambda x: x["right_of_amount"])

    if cols["account"] is None:
        cols["account"] = take(lambda x: x["org"] >= 2 and x["avg"] < 60)
    if cols["counterparty"] is None:
        cols["counterparty"] = take(lambda x: x["org"] >= 1)
    if cols["counterparty"] is None:
        cols["counterparty"] = take(lambda x: x["unique"] >= 2 and 4 <= x["avg"] <= 80)
    if cols["title"] is None:
        cols["title"] = take(lambda x: x["avg"] >= 12)
    if cols["title"] is None:
        cols["title"] = take(lambda x: x["unique"] >= 1)
    if cols["note"] is None and text_rank:
        cols["note"] = take(lambda x: x["avg"] >= 4)

    return cols


def _rows_from_xlsx(content: bytes) -> list[tuple[str, list[tuple]]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            out.append((name, rows))
    return out


def _rows_from_csv(content: bytes) -> list[tuple[str, list[tuple]]]:
    text = content.decode("utf-8-sig", errors="replace")
    if text.count("�") > 20:
        text = content.decode("cp1251", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4000], delimiters=";,\t")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";" if text.count(";") > text.count(",") else ","
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [tuple(r) for r in reader]
    return [("csv", rows)] if rows else []


def _load_sheets(content: bytes, filename: str = "") -> list[tuple[str, list[tuple]]]:
    name = (filename or "").lower()
    if name.endswith(".csv") or (
        not name.endswith((".xlsx", ".xlsm")) and (b"," in content[:200] or b";" in content[:200])
    ):
        return _rows_from_csv(content)
    return _rows_from_xlsx(content)


def _header_label(headers: list[str], idx: int | None) -> str | None:
    if idx is None or idx >= len(headers):
        return None
    h = str(headers[idx]).strip()
    return h or f"col{idx + 1}"


def inspect_payments_file(content: bytes, filename: str = "") -> dict:
    sheets = _load_sheets(content, filename)
    result = {"sheets": []}
    for sheet_name, rows in sheets:
        if not rows:
            continue
        header_idx, headers = _detect_header(rows)
        max_w = max(len(headers), max((len(r) for r in rows[:50]), default=0))
        while len(headers) < max_w:
            headers.append(f"col{len(headers)}")
        amount, debit, credit = _resolve_amount_cols(headers, rows, header_idx if header_idx >= 0 else -1)
        if header_idx < 0 and amount is None and debit is None and credit is None:
            scored = [(_score_amount_col(rows, c, 0), c) for c in range(max_w)]
            scored.sort(reverse=True)
            if scored and scored[0][0] >= 2:
                amount = scored[0][1]
        reserved = {c for c in (amount, debit, credit) if c is not None}
        cols = _infer_cols(
            headers, rows, header_idx if header_idx >= 0 else -1, reserved, amount_col=amount or debit
        )
        data_start = 0 if header_idx < 0 else header_idx + 1
        sample = []
        for row in rows[data_start : data_start + 3]:
            if not row:
                continue
            sample.append([("" if c is None else str(c)[:80]) for c in row[:12]])
        result["sheets"].append(
            {
                "name": sheet_name,
                "headerRow": None if header_idx < 0 else header_idx + 1,
                "headers": [h for h in headers if str(h).strip()],
                "mapped": {
                    "amount": _header_label(headers, amount),
                    "debit": _header_label(headers, debit),
                    "credit": _header_label(headers, credit),
                    "date": _header_label(headers, cols["date"]),
                    "counterparty": _header_label(headers, cols["counterparty"]),
                    "purpose": _header_label(headers, cols["title"]),
                    "accountOrOrg": _header_label(headers, cols["account"]),
                    "note": _header_label(headers, cols["note"]),
                    "status": _header_label(headers, cols["status"]),
                },
                "sampleRows": sample,
            }
        )
    return result


def parse_payments_file(content: bytes, filename: str = "", source: str = "excel") -> list[Payment]:
    sheets = _load_sheets(content, filename)
    errors = []
    for sheet_name, rows in sheets:
        try:
            payments = _parse_sheet(rows, source=source)
            if payments:
                return payments
            errors.append(f"лист «{sheet_name}»: строк с суммой не найдено")
        except ValueError as e:
            errors.append(f"лист «{sheet_name}»: {e}")
    raise ValueError(" | ".join(errors) if errors else "Файл пустой")


def parse_payments_xlsx(content: bytes, source: str = "excel") -> list[Payment]:
    return parse_payments_file(content, filename="file.xlsx", source=source)


def _cell(row: list, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_sheet(rows: list[tuple], source: str) -> list[Payment]:
    if not rows:
        return []

    header_idx, headers = _detect_header(rows)
    data_start = 0 if header_idx < 0 else header_idx + 1
    max_w = max(len(headers), max((len(r) for r in rows[:50]), default=0))
    while len(headers) < max_w:
        headers.append(f"col{len(headers)}")

    amount_col, debit_col, credit_col = _resolve_amount_cols(headers, rows, max(header_idx, 0) if header_idx >= 0 else -1)
    # When no header, resolve amounts against all rows from start
    if header_idx < 0 and amount_col is None and debit_col is None and credit_col is None:
        width = max_w
        scored = [(_score_amount_col(rows, c, 0), c) for c in range(width)]
        scored.sort(reverse=True)
        if scored and scored[0][0] >= 2:
            amount_col = scored[0][1]

    reserved = {c for c in (amount_col, debit_col, credit_col) if c is not None}
    infer_from = header_idx if header_idx >= 0 else -1
    cols = _infer_cols(headers, rows, infer_from, reserved, amount_col=amount_col or debit_col)

    if amount_col is None and debit_col is None and credit_col is None:
        shown = ", ".join(h for h in headers if str(h).strip())[:160]
        raise ValueError(
            f"Не нашёл колонку суммы. Заголовки в файле: [{shown}]. "
            f"Переименуй колонку в «Сумма» / «Сумма к оплате» или используй «Дебет»/«Кредит»."
        )

    out: list[Payment] = []
    for row in rows[data_start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        row = list(row)
        while len(row) < max_w:
            row.append(None)

        amount = None
        if amount_col is not None:
            amount = _parse_amount(row[amount_col])
        if amount is None and debit_col is not None:
            amount = _parse_amount(row[debit_col])
        if amount is None and credit_col is not None:
            amount = _parse_amount(row[credit_col])
        if amount is None or amount == 0:
            continue

        purpose = ""
        if cols["title"] is not None and row[cols["title"]] is not None:
            purpose = str(row[cols["title"]]).strip()

        counterparty = None
        if cols["counterparty"] is not None and row[cols["counterparty"]] is not None:
            counterparty = str(row[cols["counterparty"]]).strip() or None

        probe = _norm(f"{purpose} {counterparty or ''}")
        if probe in {"итого", "всего", "total", "сумма"}:
            continue

        account_name = None
        if cols["account"] is not None and row[cols["account"]] is not None:
            account_name = str(row[cols["account"]]).strip() or None

        note = None
        if cols["note"] is not None and row[cols["note"]] is not None:
            note = str(row[cols["note"]]).strip() or None

        # Build readable title: кому · за что
        if counterparty and purpose and purpose != counterparty:
            title = f"{counterparty} · {purpose}"
        else:
            title = purpose or counterparty or "Платёж"

        status_raw = ""
        if cols["status"] is not None and row[cols["status"]] is not None:
            status_raw = _norm(row[cols["status"]])
        status = "plan"
        if any(x in status_raw for x in ("оплач", "paid", "done", "закрыт", "исполн", "да", "true")):
            status = "done"
        elif any(x in status_raw for x in ("нов", "new")):
            status = "new"
        elif any(x in status_raw for x in ("к оплате",)):
            status = "plan"
        elif any(x in status_raw for x in ("соглас", "утверж", "ок")):
            status = "ok"

        pay_date = None
        if cols["date"] is not None:
            pay_date = _parse_date(row[cols["date"]])

        out.append(
            Payment(
                title=title[:512],
                counterparty=(counterparty[:512] if counterparty else None),
                amount=amount,
                pay_date=pay_date,
                account_name=account_name,
                status=status,
                source=source,
                note=note,
            )
        )
    return out
