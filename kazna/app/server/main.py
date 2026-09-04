from __future__ import annotations

import os
from datetime import date, datetime, timedelta
from pathlib import Path

from fastapi import Body, Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from passlib.context import CryptContext
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select, text
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.trustedhost import TrustedHostMiddleware

from .database import UPLOADS, Base, SessionLocal, engine, get_db
from .excel_import import inspect_payments_file, parse_payments_file
from .iiko_client import IikoClient, IikoError
from .iiko_sync import get_or_create_settings, settings_public, sync_iiko_invoices
from .models import Account, Payment, RequestItem, User, UserAccount
from .security import (
    UPLOAD_MAX_BYTES,
    LoginRateLimiter,
    SecurityHeadersMiddleware,
    allow_demo,
    audit,
    client_ip,
    docs_enabled,
    https_only_cookies,
    is_production,
    open_secret,
    resolve_secret,
    seal_secret,
    session_max_age,
)

ROOT = Path(__file__).resolve().parents[1]
pwd = CryptContext(schemes=["bcrypt"], deprecated="auto")

ROLE_LABELS = {
    "fin_director": "Финансовый директор",
    "manager": "Управляющий",
    "accountant": "Бухгалтер",
}
ALLOWED_ROLES = set(ROLE_LABELS)

APP_SECRET = resolve_secret()
login_limiter = LoginRateLimiter()

app = FastAPI(
    title="Казна",
    docs_url="/docs" if docs_enabled() else None,
    redoc_url="/redoc" if docs_enabled() else None,
    openapi_url="/openapi.json" if docs_enabled() else None,
)

_hosts = [h.strip() for h in (os.environ.get("KAZNA_HOSTS") or "").split(",") if h.strip()]
if _hosts:
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=_hosts)

app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=APP_SECRET,
    session_cookie="kazna_session",
    same_site="lax",
    https_only=https_only_cookies(),
    max_age=session_max_age(),
)


def _migrate_schema() -> None:
    """Add new columns on existing DBs (create_all alone won't alter)."""
    stmts = [
        "ALTER TABLE requests ADD COLUMN comment TEXT",
        "ALTER TABLE requests ADD COLUMN due_date DATE",
        "ALTER TABLE requests ADD COLUMN decided_by INTEGER",
        "ALTER TABLE requests ADD COLUMN payment_id INTEGER",
        "ALTER TABLE requests ADD COLUMN created_at TIMESTAMP",
        "ALTER TABLE requests ADD COLUMN updated_at TIMESTAMP",
        "ALTER TABLE payments ADD COLUMN external_id VARCHAR(255)",
        "ALTER TABLE accounts ADD COLUMN site VARCHAR(255)",
        "ALTER TABLE accounts ADD COLUMN iiko_org_id VARCHAR(64)",
    ]
    with engine.begin() as conn:
        for sql in stmts:
            try:
                conn.execute(text(sql))
            except Exception:
                pass


def init_db() -> None:
    Base.metadata.create_all(bind=engine)
    _migrate_schema()
    if not allow_demo():
        return
    db = SessionLocal()
    try:
        if db.scalar(select(func.count()).select_from(User)) == 0:
            db.add_all(
                [
                    User(
                        email="fin@kazna.local",
                        password_hash=pwd.hash("kazna2026"),
                        name="Анна Финдир",
                        role="fin_director",
                        role_label="Финансовый директор",
                    ),
                    User(
                        email="manager@kazna.local",
                        password_hash=pwd.hash("kazna2026"),
                        name="Игорь Управляющий",
                        role="manager",
                        role_label="Управляющий · Невский",
                        site="Невский",
                    ),
                ]
            )
            db.commit()
            audit("demo_users_seeded")
    finally:
        db.close()


@app.on_event("startup")
def on_startup() -> None:
    init_db()


def current_user(request: Request, db: Session) -> User | None:
    uid = request.session.get("uid")
    if not uid:
        return None
    return db.get(User, uid)


def require_user(request: Request, db: Session = Depends(get_db)) -> User:
    user = current_user(request, db)
    if not user or not user.active:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return user


def require_fin(user: User = Depends(require_user)) -> User:
    if user.role != "fin_director":
        raise HTTPException(status_code=403, detail="Только финдир")
    return user


def require_office(user: User = Depends(require_user)) -> User:
    """Финдир или бухгалтер — кабинет сети, не точка управляющего."""
    if user.role not in {"fin_director", "accountant"}:
        raise HTTPException(status_code=403, detail="Нет доступа")
    return user


class LoginIn(BaseModel):
    email: str
    password: str


class RequestCreate(BaseModel):
    title: str = Field(min_length=2, max_length=512)
    amount: float = Field(gt=0)
    due_date: str | None = None
    comment: str | None = None
    meta: str | None = None
    priority: str = "normal"


class RequestPatch(BaseModel):
    status: str | None = None
    priority: str | None = None
    comment: str | None = None
    due_date: str | None = None
    to_calendar: bool = False


class PersonCreate(BaseModel):
    name: str = Field(min_length=2, max_length=255)
    email: str = Field(min_length=3, max_length=255)
    password: str = Field(min_length=8, max_length=128)
    role: str = "manager"
    site: str | None = None
    account_ids: list[int] = Field(default_factory=list)
    active: bool = True


class PersonPatch(BaseModel):
    name: str | None = None
    email: str | None = None
    password: str | None = None
    role: str | None = None
    site: str | None = None
    account_ids: list[int] | None = None
    active: bool | None = None


class AccountCreate(BaseModel):
    name: str = Field(min_length=1, max_length=255)
    kind: str = "р/с"
    org: str | None = None
    site: str | None = None
    balance: float = 0


class AccountPatch(BaseModel):
    name: str | None = None
    kind: str | None = None
    org: str | None = None
    site: str | None = None
    balance: float | None = None


class PaymentPatch(BaseModel):
    title: str | None = None
    counterparty: str | None = None
    amount: float | None = None
    pay_date: str | None = None  # ISO or empty to clear
    clear_date: bool = False
    postpone_days: int | None = None  # shift from current/today
    account_name: str | None = None
    status: str | None = None
    note: str | None = None


class PaymentCreate(BaseModel):
    title: str = Field(min_length=1, max_length=512)
    counterparty: str | None = None
    amount: float = Field(gt=0)
    pay_date: str | None = None
    account_name: str | None = None
    status: str = "plan"
    note: str | None = None


def fmt_money(n: float) -> str:
    n = float(n or 0)
    if abs(n) >= 1_000_000:
        return f"{n / 1_000_000:.1f}".replace(".", ",") + " млн"
    if abs(n) >= 1_000:
        return f"{round(n / 1000):,}".replace(",", " ") + " тыс."
    return f"{round(n):,}".replace(",", " ") + " ₽"


def parse_iso_date(s: str | None) -> date | None:
    if not s:
        return None
    s = str(s).strip()
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        for fmt in ("%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
    return None


def request_dict(r: RequestItem, creator: User | None = None) -> dict:
    return {
        "id": r.id,
        "title": r.title,
        "amount": r.amount,
        "amountLabel": fmt_money(r.amount),
        "meta": r.meta,
        "comment": r.comment,
        "status": r.status,
        "priority": r.priority,
        "site": r.site,
        "dueDate": r.due_date.isoformat() if r.due_date else None,
        "dueDateLabel": r.due_date.strftime("%d.%m.%Y") if r.due_date else "без срока",
        "createdBy": r.created_by,
        "creatorName": creator.name if creator else None,
        "paymentId": r.payment_id,
        "createdAt": r.created_at.isoformat() if r.created_at else None,
    }


def payment_dict(p: Payment) -> dict:
    # title often "контрагент · назначение" — split for UI
    purpose = p.title or ""
    counterparty = p.counterparty
    if counterparty and purpose.startswith(counterparty):
        rest = purpose[len(counterparty) :].lstrip(" ·")
        if rest:
            purpose = rest
    return {
        "id": p.id,
        "title": p.title,
        "counterparty": counterparty,
        "purpose": purpose if purpose != (counterparty or "") else None,
        "amount": p.amount,
        "amountLabel": fmt_money(p.amount),
        "date": p.pay_date.isoformat() if p.pay_date else None,
        "dateLabel": p.pay_date.strftime("%d.%m.%Y") if p.pay_date else "без даты",
        "status": p.status,
        "source": p.source,
        "account": p.account_name,
        "note": p.note,
        "externalId": p.external_id,
    }


@app.get("/api/health")
def health():
    return {
        "ok": True,
        "service": "kazna",
        "production": is_production(),
        "httpsCookies": https_only_cookies(),
    }


@app.get("/api/public/bootstrap")
def public_bootstrap():
    """Safe public flags for the login page (no secrets)."""
    return {
        "demoMode": allow_demo(),
        "production": is_production(),
    }


@app.post("/api/login")
def login(payload: LoginIn, request: Request, db: Session = Depends(get_db)):
    email = str(payload.email).strip().lower()
    ip = client_ip(request)
    try:
        login_limiter.check(ip, email)
    except PermissionError as e:
        audit("login_rate_limited", email=email, ip=ip)
        raise HTTPException(status_code=429, detail=str(e)) from e

    user = db.scalar(select(User).where(User.email == email))
    if not user or not pwd.verify(payload.password, user.password_hash):
        login_limiter.fail(ip, email)
        audit("login_failed", email=email, ip=ip)
        raise HTTPException(status_code=400, detail="Неверный логин или пароль")
    if not user.active:
        audit("login_inactive", email=email, ip=ip)
        raise HTTPException(status_code=403, detail="Учётная запись отключена")

    login_limiter.success(ip, email)
    request.session.clear()
    request.session["uid"] = user.id
    audit("login_ok", email=email, ip=ip, role=user.role)
    return {
        "ok": True,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role,
            "roleLabel": user.role_label,
            "site": user.site,
        },
    }


@app.post("/api/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/api/me")
def me(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or not user.active:
        raise HTTPException(status_code=401, detail="Unauthorized")
    aids = _account_ids_for_user(db, user.id)
    accounts = []
    if aids:
        accounts = [
            {"id": a.id, "name": a.name, "org": a.org}
            for a in db.scalars(select(Account).where(Account.id.in_(aids)).order_by(Account.id)).all()
        ]
    return {
        "id": user.id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "roleLabel": user.role_label,
        "site": user.site,
        "accountIds": aids,
        "accounts": accounts,
    }


def _is_cash_desk(kind: str | None) -> bool:
    return "касс" in (kind or "").lower()


def _org_key(account: Account) -> str:
    return (account.org or account.name or "Без организации").strip() or "Без организации"


def _payment_matches_org(p: Payment, org: str, account_names: set[str]) -> bool:
    an = (p.account_name or "").strip()
    if an and (an in account_names or an == org):
        return True
    blob = f"{p.title or ''} {p.counterparty or ''} {an}".lower()
    return bool(org) and org.lower() in blob


@app.get("/api/overview")
def overview(user: User = Depends(require_office), db: Session = Depends(get_db)):
    """Утренний кассовый разрыв: по организациям + кассы точек."""
    today = date.today()
    accounts = list(db.scalars(select(Account).order_by(Account.id)).all())
    settlement = [a for a in accounts if not _is_cash_desk(a.kind)]
    cash_desks = [a for a in accounts if _is_cash_desk(a.kind)]

    settlement_total = sum(a.balance for a in settlement)
    cash_total = sum(a.balance for a in cash_desks)
    accounts_total = settlement_total + cash_total

    open_pays = list(
        db.scalars(
            select(Payment)
            .where(Payment.status != "done")
            .order_by(Payment.pay_date.is_(None), Payment.pay_date, Payment.id)
        ).all()
    )
    due_today = [
        p
        for p in open_pays
        if p.pay_date is None or p.pay_date <= today
    ]
    week_pays = [
        p
        for p in open_pays
        if p.pay_date is not None and today < p.pay_date <= today + timedelta(days=7)
    ]

    pay_today_sum = sum(p.amount for p in due_today)
    after = settlement_total - pay_today_sum if settlement else None
    gap = max(0.0, -(after or 0)) if after is not None else None

    # Group settlement accounts by org
    by_org: dict[str, list[Account]] = {}
    for a in settlement:
        by_org.setdefault(_org_key(a), []).append(a)

    org_boards = []
    assigned_ids: set[int] = set()
    for org, accs in sorted(by_org.items(), key=lambda x: x[0].lower()):
        names = {a.name for a in accs} | {org}
        bal = sum(a.balance for a in accs)
        matched = [p for p in due_today if _payment_matches_org(p, org, names)]
        for p in matched:
            assigned_ids.add(p.id)
        due_sum = sum(p.amount for p in matched)
        after_org = bal - due_sum
        gap_org = max(0.0, -after_org)
        org_boards.append(
            {
                "org": org,
                "balance": bal,
                "balanceLabel": fmt_money(bal),
                "dueToday": due_sum,
                "dueTodayLabel": fmt_money(due_sum),
                "after": after_org,
                "afterLabel": fmt_money(after_org),
                "gap": gap_org,
                "gapLabel": fmt_money(gap_org),
                "hasGap": after_org < 0,
                "canPayAll": due_sum > 0 and after_org >= 0,
                "accounts": [account_dict(a) for a in accs],
                "payments": [payment_dict(p) for p in matched[:12]],
                "paymentCount": len(matched),
            }
        )

    unassigned = [p for p in due_today if p.id not in assigned_ids]
    unassigned_sum = sum(p.amount for p in unassigned)

    pending = db.scalar(
        select(func.count())
        .select_from(RequestItem)
        .where(RequestItem.status.in_(("new", "reviewing")))
    )
    requests = db.scalars(select(RequestItem).order_by(RequestItem.id.desc()).limit(20)).all()

    if after is not None and after < 0:
        verdict = f"Кассовый разрыв {fmt_money(gap or 0)} — сегодня всё не закрыть без переноса."
        verdict_kind = "danger"
    elif pay_today_sum <= 0:
        verdict = "На сегодня обязательных оплат нет — можно планировать неделю."
        verdict_kind = "ok"
    elif after is not None:
        verdict = f"Можно оплатить сегодня: останется {fmt_money(after)} на р/с."
        verdict_kind = "ok"
    else:
        verdict = "Загрузите остатки р/с, чтобы видеть разрыв по организациям."
        verdict_kind = "warn"

    return {
        "today": today.isoformat(),
        "verdict": verdict,
        "verdictKind": verdict_kind,
        "accountsTotal": accounts_total,
        "accountsTotalLabel": fmt_money(accounts_total) if accounts else "нет данных",
        "settlementTotal": settlement_total,
        "settlementTotalLabel": fmt_money(settlement_total) if settlement else "нет данных",
        "cashTotal": cash_total,
        "cashTotalLabel": fmt_money(cash_total) if cash_desks else "нет касс",
        "payToday": pay_today_sum,
        "payTodayLabel": fmt_money(pay_today_sum),
        "weekAhead": sum(p.amount for p in week_pays),
        "weekAheadLabel": fmt_money(sum(p.amount for p in week_pays)),
        "afterPay": after,
        "afterPayLabel": fmt_money(after) if after is not None else "—",
        "gap": gap,
        "gapLabel": fmt_money(gap) if gap is not None else "—",
        "hasGap": bool(after is not None and after < 0),
        "hasAccounts": bool(settlement),
        "hasPayments": db.scalar(select(func.count()).select_from(Payment)) > 0,
        "pendingRequests": int(pending or 0),
        "organizations": org_boards,
        "cashDesks": [
            {
                **account_dict(a),
                "siteLabel": a.site or "точка не указана",
            }
            for a in sorted(cash_desks, key=lambda x: ((x.site or ""), x.name))
        ],
        "unassigned": {
            "dueToday": unassigned_sum,
            "dueTodayLabel": fmt_money(unassigned_sum),
            "paymentCount": len(unassigned),
            "payments": [payment_dict(p) for p in unassigned[:12]],
        },
        "accounts": [account_dict(a) for a in accounts],
        "payments": [payment_dict(p) for p in due_today[:30]],
        "weekPayments": [payment_dict(p) for p in week_pays[:20]],
        "requests": [request_dict(r) for r in requests],
    }


@app.get("/api/payments")
def list_payments(
    source: str | None = None,
    account: str | None = None,
    status: str | None = None,
    user: User = Depends(require_office),
    db: Session = Depends(get_db),
):
    """Payments calendar — office only (managers use /api/requests)."""
    q = select(Payment).order_by(Payment.pay_date.is_(None), Payment.pay_date, Payment.id)
    if source:
        q = q.where(Payment.source == source)
    if account:
        q = q.where(Payment.account_name == account)
    if status:
        q = q.where(Payment.status == status)
    rows = db.scalars(q).all()
    return [payment_dict(p) for p in rows]


@app.post("/api/payments")
def create_payment(payload: PaymentCreate, user: User = Depends(require_office), db: Session = Depends(get_db)):
    status = payload.status if payload.status in {"new", "plan", "ok", "done"} else "plan"
    title = payload.title.strip()
    counterparty = (payload.counterparty or "").strip() or None
    if counterparty and counterparty not in title:
        title = f"{counterparty} · {title}"
    pay = Payment(
        title=title[:512],
        counterparty=counterparty,
        amount=float(payload.amount),
        pay_date=parse_iso_date(payload.pay_date),
        account_name=(payload.account_name or "").strip() or None,
        status=status,
        source="manual",
        note=(payload.note or "").strip() or None,
    )
    db.add(pay)
    db.commit()
    db.refresh(pay)
    return payment_dict(pay)


@app.patch("/api/payments/{payment_id}")
def patch_payment(
    payment_id: int,
    payload: PaymentPatch,
    user: User = Depends(require_office),
    db: Session = Depends(get_db),
):
    pay = db.get(Payment, payment_id)
    if not pay:
        raise HTTPException(404, "Платёж не найден")

    if payload.title is not None:
        pay.title = payload.title.strip()[:512] or pay.title
    if payload.counterparty is not None:
        pay.counterparty = payload.counterparty.strip() or None
    if payload.amount is not None:
        if payload.amount <= 0:
            raise HTTPException(400, "Сумма должна быть больше 0")
        pay.amount = float(payload.amount)
    if payload.clear_date:
        pay.pay_date = None
    elif payload.postpone_days is not None:
        days = int(payload.postpone_days)
        base = pay.pay_date or date.today()
        pay.pay_date = base + timedelta(days=days)
        if pay.status == "done":
            pay.status = "plan"
    elif payload.pay_date is not None:
        pay.pay_date = parse_iso_date(payload.pay_date) if payload.pay_date.strip() else None
    if payload.account_name is not None:
        pay.account_name = payload.account_name.strip() or None
    if payload.status is not None:
        if payload.status not in {"new", "plan", "ok", "done"}:
            raise HTTPException(400, "Неверный статус")
        pay.status = payload.status
    if payload.note is not None:
        pay.note = payload.note.strip() or None

    db.commit()
    db.refresh(pay)
    return payment_dict(pay)


@app.delete("/api/payments/source/{source_name}")
def delete_payments_by_source(
    source_name: str,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    """Remove noisy Excel imports without touching iiko/manual."""
    if source_name not in {"excel", "iiko"}:
        raise HTTPException(400, "Можно удалить только excel или iiko")
    n = db.scalar(select(func.count()).select_from(Payment).where(Payment.source == source_name)) or 0
    db.execute(delete(Payment).where(Payment.source == source_name))
    db.commit()
    return {"ok": True, "deleted": int(n), "message": f"Удалено платежей ({source_name}): {n}"}


def _import_payments(
    content: bytes,
    filename: str,
    source: str,
    replace: bool,
    db: Session,
    *,
    create_accounts: bool = False,
) -> dict:
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in (filename or "upload.xlsx"))
    dest = UPLOADS / f"{stamp}_{safe}"
    dest.write_bytes(content)

    try:
        payments, report = parse_payments_file(content, filename=filename, source=source)
    except Exception as e:
        raise HTTPException(400, f"Не разобрал файл: {e}") from e

    if not payments:
        raise HTTPException(400, "В файле не нашёл ни одной строки с суммой")

    if replace:
        db.execute(delete(Payment).where(Payment.source == source))

    db.add_all(payments)
    db.commit()

    org_names = sorted({p.account_name for p in payments if p.account_name})
    # Do NOT invent р/с from Excel payment columns — that made org/account chaos.
    # Accounts come from: Excel остатки, ручное создание, iiko org names.
    if create_accounts and org_names:
        existing = {a.name for a in db.scalars(select(Account)).all()}
        for name in org_names:
            if name not in existing:
                kind = "ИП" if name.upper().startswith("ИП") else "юрлицо"
                db.add(Account(name=name, kind=kind, balance=0, org=name))
        db.commit()

    no_date = sum(1 for p in payments if not p.pay_date)
    no_who = sum(1 for p in payments if not p.counterparty)
    warnings = list(report.get("warnings") or [])
    if no_who > len(payments) // 2:
        warnings.append("у многих строк не найден контрагент (кому)")
    if no_date > len(payments) // 2:
        warnings.append("у многих строк нет даты/срока")

    sheet_bits = [
        f"«{s['name']}»:{s['imported']}" + (f" ({s['orgHint']})" if s.get("orgHint") else "")
        for s in report.get("sheets", [])
        if s.get("imported")
    ]
    sample = [payment_dict(p) for p in payments[:3]]
    msg = (
        f"Загружено платежей: {len(payments)} ({source}) "
        f"с {report.get('sheetsWithData', 0)} из {report.get('sheetsTotal', 0)} листов"
    )
    if sheet_bits:
        msg += ": " + "; ".join(sheet_bits)
    soft = [w for w in warnings if "строк с суммой не найдено" not in w]
    if soft:
        msg += ". Внимание: " + "; ".join(soft[:4])

    return {
        "ok": True,
        "imported": len(payments),
        "file": dest.name,
        "source": source,
        "message": msg,
        "warnings": warnings,
        "sheets": report.get("sheets", []),
        "orgs": org_names,
        "sample": sample,
    }


@app.post("/api/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    replace: bool = True,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(400, "Нужен файл .xlsx или .csv")
    content = await file.read()
    if len(content) > UPLOAD_MAX_BYTES:
        raise HTTPException(400, "Файл больше 15 МБ")
    audit("import_excel", user=user.email, bytes=len(content), name=file.filename)
    return _import_payments(content, file.filename or "payments.xlsx", "excel", replace, db)


@app.post("/api/import/iiko")
async def import_iiko(
    file: UploadFile = File(...),
    replace: bool = True,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    """Import iiko export (xlsx/csv). Live API credentials come next."""
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(400, "Нужен выгрузка iiko: .xlsx или .csv")
    content = await file.read()
    if len(content) > UPLOAD_MAX_BYTES:
        raise HTTPException(400, "Файл больше 15 МБ")
    audit("import_iiko_file", user=user.email, bytes=len(content), name=file.filename)
    return _import_payments(content, file.filename or "iiko.xlsx", "iiko", replace, db)


@app.post("/api/import/inspect")
async def import_inspect(
    file: UploadFile = File(...),
    user: User = Depends(require_fin),
):
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(400, "Нужен файл .xlsx или .csv")
    content = await file.read()
    if len(content) > UPLOAD_MAX_BYTES:
        raise HTTPException(400, "Файл больше 15 МБ")
    try:
        info = inspect_payments_file(content, filename=file.filename or "")
    except Exception as e:
        raise HTTPException(400, f"Не открыл файл: {e}") from e
    return {"ok": True, "filename": file.filename, **info}


class IikoConfigIn(BaseModel):
    api_login: str | None = None
    organization_id: str | None = None
    org_id: str | None = None  # alias
    days_back: int | None = 30
    note: str | None = None
    enabled: bool = True


class IikoSyncIn(BaseModel):
    days_back: int | None = None
    organization_id: str | None = None
    replace: bool = False


def _iiko_login(row) -> str:
    env_login = (os.environ.get("IIKO_API_LOGIN") or "").strip()
    if env_login:
        return env_login
    try:
        return open_secret(row.api_login or "", APP_SECRET)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e


@app.get("/api/import/iiko/config")
def iiko_config_get(user: User = Depends(require_fin), db: Session = Depends(get_db)):
    return settings_public(get_or_create_settings(db))


@app.post("/api/import/iiko/config")
def iiko_config_save(
    payload: IikoConfigIn,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    row = get_or_create_settings(db)
    if payload.api_login is not None and payload.api_login.strip():
        row.api_login = seal_secret(payload.api_login.strip(), APP_SECRET)
        audit("iiko_login_saved", user=user.email)
    org = payload.organization_id or payload.org_id
    if org is not None:
        row.organization_id = org.strip() or None
        if not row.organization_id:
            row.organization_name = None
    if payload.days_back is not None:
        row.days_back = max(1, min(int(payload.days_back), 366))
    if payload.note is not None:
        row.note = payload.note.strip() or None
    row.enabled = bool(payload.enabled)
    db.commit()
    db.refresh(row)
    return {"ok": True, "message": "Настройки iiko сохранены.", **settings_public(row)}


@app.post("/api/import/iiko/test")
def iiko_test(user: User = Depends(require_fin), db: Session = Depends(get_db)):
    row = get_or_create_settings(db)
    login = _iiko_login(row)
    if not login:
        raise HTTPException(400, "Сначала сохраните apiLogin или задайте IIKO_API_LOGIN")
    try:
        client = IikoClient(login)
        client.auth()
        orgs = client.organizations()
    except IikoError as e:
        row.last_error = str(e)
        db.commit()
        raise HTTPException(400, str(e)) from e
    row.last_error = None
    db.commit()
    return {
        "ok": True,
        "message": f"Связь есть. Организаций: {len(orgs)}",
        "organizations": orgs,
    }


@app.get("/api/import/iiko/organizations")
def iiko_organizations(user: User = Depends(require_fin), db: Session = Depends(get_db)):
    row = get_or_create_settings(db)
    login = _iiko_login(row)
    if not login:
        raise HTTPException(400, "Сначала сохраните apiLogin или задайте IIKO_API_LOGIN")
    try:
        client = IikoClient(login)
        client.auth()
        return {"organizations": client.organizations()}
    except IikoError as e:
        raise HTTPException(400, str(e)) from e


@app.post("/api/import/iiko/sync")
def iiko_sync(
    payload: IikoSyncIn = Body(default_factory=IikoSyncIn),
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    row = get_or_create_settings(db)
    login = _iiko_login(row)
    try:
        result = sync_iiko_invoices(
            db,
            api_login=login or None,
            organization_id=payload.organization_id or row.organization_id,
            days_back=payload.days_back,
            replace=payload.replace,
        )
    except IikoError as e:
        row.last_error = str(e)
        db.commit()
        raise HTTPException(400, str(e)) from e
    except Exception as e:
        row.last_error = str(e)
        db.commit()
        raise HTTPException(400, f"Синк iiko не удался: {e}") from e
    audit("iiko_sync", user=user.email, imported=result.get("imported"))
    return result


@app.post("/api/accounts/set-balances")
async def set_balances(
    file: UploadFile = File(...),
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    """Upsert balances by account name — does not wipe iiko/manual accounts."""
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(400, "Нужен файл .xlsx или .csv")
    content = await file.read()
    if len(content) > UPLOAD_MAX_BYTES:
        raise HTTPException(400, "Файл больше 15 МБ")
    try:
        rows = parse_balance_sheet(content)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не разобрал остатки: {e}") from e

    existing = {a.name: a for a in db.scalars(select(Account)).all()}
    updated = 0
    created = 0
    for row in rows:
        name = row["name"]
        if name in existing:
            acc = existing[name]
            acc.balance = row["balance"]
            if row.get("org"):
                acc.org = row["org"]
            acc.kind = row.get("kind") or acc.kind or "р/с"
            updated += 1
        else:
            db.add(
                Account(
                    name=name,
                    kind=row.get("kind") or "р/с",
                    balance=row["balance"],
                    org=row.get("org") or name,
                )
            )
            created += 1
    db.commit()
    return {
        "ok": True,
        "accounts": len(rows),
        "updated": updated,
        "created": created,
        "message": f"Остатки: обновлено {updated}, новых счетов {created}",
    }


def parse_balance_sheet(content: bytes) -> list[dict]:
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    name_i = None
    bal_i = None
    for i, h in enumerate(headers):
        hn = h.strip().lower().replace("ё", "е")
        if name_i is None and any(x in hn for x in ("счёт", "счет", "организ", "account", "название")):
            name_i = i
        if bal_i is None and any(x in hn for x in ("остаток", "баланс", "balance", "сумма")):
            bal_i = i
    if name_i is None or bal_i is None:
        raise HTTPException(400, "Нужны колонки: Счёт/Организация и Остаток")
    out = []
    for row in rows[1:]:
        if not row or row[name_i] is None:
            continue
        try:
            bal = float(str(row[bal_i]).replace(" ", "").replace(",", ".").replace("\u00a0", ""))
        except Exception:
            continue
        out.append({"name": str(row[name_i]).strip(), "balance": bal, "kind": "р/с", "org": str(row[name_i]).strip()})
    if not out:
        raise HTTPException(400, "Пустой файл остатков")
    return out


def _account_ids_for_user(db: Session, user_id: int) -> list[int]:
    return list(
        db.scalars(select(UserAccount.account_id).where(UserAccount.user_id == user_id)).all()
    )


def _set_user_accounts(db: Session, user_id: int, account_ids: list[int]) -> None:
    db.execute(delete(UserAccount).where(UserAccount.user_id == user_id))
    if not account_ids:
        return
    valid = set(db.scalars(select(Account.id).where(Account.id.in_(account_ids))).all())
    for aid in account_ids:
        if aid in valid:
            db.add(UserAccount(user_id=user_id, account_id=aid))


def _role_label(role: str, site: str | None = None) -> str:
    base = ROLE_LABELS.get(role, role)
    if role == "manager" and site:
        return f"{base} · {site}"
    return base


def person_dict(u: User, db: Session) -> dict:
    aids = _account_ids_for_user(db, u.id)
    accounts = []
    if aids:
        accounts = [
            {"id": a.id, "name": a.name, "org": a.org}
            for a in db.scalars(select(Account).where(Account.id.in_(aids)).order_by(Account.id)).all()
        ]
    return {
        "id": u.id,
        "name": u.name,
        "email": u.email,
        "role": u.role,
        "roleLabel": u.role_label or _role_label(u.role, u.site),
        "site": u.site,
        "siteLabel": u.site or ("все точки" if u.role == "fin_director" else "не назначена"),
        "active": u.active,
        "accountIds": aids,
        "accounts": accounts,
    }


def account_dict(a: Account) -> dict:
    return {
        "id": a.id,
        "name": a.name,
        "kind": a.kind,
        "balance": a.balance,
        "balanceLabel": fmt_money(a.balance),
        "org": a.org,
        "site": a.site,
        "iikoOrgId": a.iiko_org_id,
        "isCashDesk": _is_cash_desk(a.kind),
    }


@app.get("/api/roles")
def list_roles(user: User = Depends(require_fin)):
    return [{"id": k, "label": v} for k, v in ROLE_LABELS.items()]


@app.get("/api/accounts")
def list_accounts(user: User = Depends(require_user), db: Session = Depends(get_db)):
    rows = db.scalars(select(Account).order_by(Account.id)).all()
    if user.role == "fin_director":
        return [account_dict(a) for a in rows]
    allowed = set(_account_ids_for_user(db, user.id))
    if not allowed:
        return []
    return [account_dict(a) for a in rows if a.id in allowed]


@app.post("/api/accounts")
def create_account(payload: AccountCreate, user: User = Depends(require_fin), db: Session = Depends(get_db)):
    kind = (payload.kind or "р/с").strip()
    site = (payload.site or "").strip() or None
    if _is_cash_desk(kind) and not site:
        raise HTTPException(400, "Для кассы укажите точку")
    if payload.org and payload.org.strip():
        org = payload.org.strip()
    elif _is_cash_desk(kind):
        org = None
    else:
        org = payload.name.strip()
    acc = Account(
        name=payload.name.strip(),
        kind=kind,
        org=org,
        site=site,
        balance=float(payload.balance or 0),
    )
    db.add(acc)
    db.commit()
    db.refresh(acc)
    return account_dict(acc)


@app.patch("/api/accounts/{account_id}")
def patch_account(
    account_id: int,
    payload: AccountPatch,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    acc = db.get(Account, account_id)
    if not acc:
        raise HTTPException(404, "Счёт не найден")
    if payload.name is not None:
        acc.name = payload.name.strip()
    if payload.kind is not None:
        acc.kind = payload.kind.strip()
    if payload.org is not None:
        acc.org = payload.org.strip() or None
    if payload.site is not None:
        acc.site = payload.site.strip() or None
    if payload.balance is not None:
        acc.balance = float(payload.balance)
    if _is_cash_desk(acc.kind) and not acc.site:
        raise HTTPException(400, "Для кассы укажите точку")
    db.commit()
    db.refresh(acc)
    return account_dict(acc)


@app.get("/api/people")
def people(user: User = Depends(require_fin), db: Session = Depends(get_db)):
    rows = db.scalars(select(User).order_by(User.id)).all()
    return [person_dict(u, db) for u in rows]


@app.post("/api/people")
def create_person(payload: PersonCreate, user: User = Depends(require_fin), db: Session = Depends(get_db)):
    email = payload.email.strip().lower()
    if db.scalar(select(User).where(User.email == email)):
        raise HTTPException(400, "Такой email уже есть")
    role = payload.role if payload.role in ALLOWED_ROLES else "manager"
    site = (payload.site or "").strip() or None
    person = User(
        email=email,
        password_hash=pwd.hash(payload.password),
        name=payload.name.strip(),
        role=role,
        role_label=_role_label(role, site),
        site=site,
        active=bool(payload.active),
    )
    db.add(person)
    db.flush()
    _set_user_accounts(db, person.id, payload.account_ids)
    db.commit()
    db.refresh(person)
    return person_dict(person, db)


@app.patch("/api/people/{person_id}")
def patch_person(
    person_id: int,
    payload: PersonPatch,
    user: User = Depends(require_fin),
    db: Session = Depends(get_db),
):
    person = db.get(User, person_id)
    if not person:
        raise HTTPException(404, "Пользователь не найден")

    if payload.email is not None:
        email = payload.email.strip().lower()
        other = db.scalar(select(User).where(User.email == email, User.id != person.id))
        if other:
            raise HTTPException(400, "Такой email уже есть")
        person.email = email
    if payload.name is not None:
        person.name = payload.name.strip()
    if payload.password:
        if len(payload.password) < 8:
            raise HTTPException(400, "Пароль минимум 8 символов")
        person.password_hash = pwd.hash(payload.password)
    if payload.role is not None:
        if payload.role not in ALLOWED_ROLES:
            raise HTTPException(400, "Неизвестная роль")
        # Don't lock yourself out of being the last fin director
        if person.id == user.id and payload.role != "fin_director":
            others = db.scalar(
                select(func.count())
                .select_from(User)
                .where(User.role == "fin_director")
                .where(User.active.is_(True))
                .where(User.id != person.id)
            )
            if not others:
                raise HTTPException(400, "Нужен хотя бы один активный финдир")
        person.role = payload.role
    if payload.site is not None:
        person.site = payload.site.strip() or None
    if payload.active is not None:
        if person.id == user.id and not payload.active:
            raise HTTPException(400, "Нельзя выключить самого себя")
        person.active = bool(payload.active)
    person.role_label = _role_label(person.role, person.site)
    if payload.account_ids is not None:
        _set_user_accounts(db, person.id, payload.account_ids)
    db.commit()
    db.refresh(person)
    return person_dict(person, db)


# ——— Requests: manager ↔ fin director ———


@app.get("/api/requests")
def list_requests(user: User = Depends(require_user), db: Session = Depends(get_db)):
    q = select(RequestItem).order_by(RequestItem.id.desc())
    if user.role == "manager":
        q = q.where(RequestItem.created_by == user.id)
    rows = db.scalars(q).all()
    creators = {u.id: u for u in db.scalars(select(User)).all()}
    return [request_dict(r, creators.get(r.created_by)) for r in rows]


@app.post("/api/requests")
def create_request(payload: RequestCreate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    if user.role not in {"manager", "fin_director"}:
        raise HTTPException(403, "Нет доступа")
    priority = payload.priority if payload.priority in {"urgent", "high", "normal", "low"} else "normal"
    item = RequestItem(
        title=payload.title.strip()[:512],
        amount=float(payload.amount),
        meta=(payload.meta or "").strip()[:512] or None,
        comment=(payload.comment or "").strip() or None,
        status="new",
        priority=priority,
        site=user.site,
        due_date=parse_iso_date(payload.due_date),
        created_by=user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return request_dict(item, user)


@app.patch("/api/requests/{request_id}")
def patch_request(
    request_id: int,
    payload: RequestPatch,
    user: User = Depends(require_user),
    db: Session = Depends(get_db),
):
    item = db.get(RequestItem, request_id)
    if not item:
        raise HTTPException(404, "Заявка не найдена")

    is_fin = user.role == "fin_director"
    is_owner = item.created_by == user.id

    if not is_fin and not is_owner:
        raise HTTPException(403, "Нет доступа")

    if payload.priority and is_fin:
        if payload.priority not in {"urgent", "high", "normal", "low"}:
            raise HTTPException(400, "Неверный приоритет")
        item.priority = payload.priority

    if payload.comment is not None and (is_fin or is_owner):
        item.comment = payload.comment.strip() or None

    if payload.due_date is not None and (is_fin or (is_owner and item.status == "new")):
        item.due_date = parse_iso_date(payload.due_date)

    if payload.status:
        if not is_fin and not (is_owner and payload.status == "new" and item.status == "new"):
            # manager can only cancel own new requests
            if is_owner and payload.status == "rejected" and item.status in {"new", "reviewing"}:
                item.status = "rejected"
            elif not is_fin:
                raise HTTPException(403, "Статус меняет только финдир")
        if is_fin:
            allowed = {"new", "reviewing", "approved", "rejected", "scheduled", "paid"}
            if payload.status not in allowed:
                raise HTTPException(400, "Неверный статус")
            item.status = payload.status
            item.decided_by = user.id

    # Approve / to calendar → create Payment linked to request
    if is_fin and (payload.to_calendar or payload.status in {"approved", "scheduled"}):
        if item.payment_id is None:
            pay = Payment(
                title=f"{item.site + ' · ' if item.site else ''}{item.title}",
                counterparty=item.site,
                amount=item.amount,
                pay_date=item.due_date or date.today(),
                account_name=None,
                status="plan",
                source="request",
                note=item.comment,
            )
            db.add(pay)
            db.flush()
            item.payment_id = pay.id
        if item.status not in {"paid", "rejected"}:
            item.status = "scheduled" if payload.to_calendar or payload.status == "scheduled" else "approved"
        item.decided_by = user.id

    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    creator = db.get(User, item.created_by) if item.created_by else None
    return request_dict(item, creator)


# Static frontend
app.mount("/css", StaticFiles(directory=ROOT / "css"), name="css")
app.mount("/js", StaticFiles(directory=ROOT / "js"), name="js")


@app.get("/")
def root():
    return FileResponse(ROOT / "index.html")


@app.get("/{page_name}.html")
def html_page(page_name: str):
    path = ROOT / f"{page_name}.html"
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)
